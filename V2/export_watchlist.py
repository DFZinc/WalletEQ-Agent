"""
export_watchlist.py
-------------------
Reads watchlist.json and exports it to a formatted Excel spreadsheet.

Run manually anytime:
    python export_watchlist.py

Or schedule it — it always overwrites watchlist_analysis.xlsx with the latest data.

Sheets:
    1. Wallet Summary  — one row per wallet, all scores and P&L metrics
    2. Trade History   — full trade history for all wallets (buys green, sells red)
    3. Recent Activity — watchlist monitoring activity log
"""

import json
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

WATCHLIST_FILE = "watchlist.json"
OUTPUT_FILE    = "watchlist_analysis.xlsx"


def load_watchlist() -> dict:
    if not os.path.exists(WATCHLIST_FILE):
        print(f"ERROR: {WATCHLIST_FILE} not found. Run the agent first.")
        sys.exit(1)
    with open(WATCHLIST_FILE, "r") as f:
        return json.load(f)


def header_cell(cell, value):
    cell.value     = value
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color="1F3864")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def body_cell(cell, value, fmt=None, fill=None, bold=False):
    cell.value     = value
    cell.font      = Font(name="Arial", bold=bold, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt


ALT_FILL    = PatternFill("solid", start_color="EEF2F7")
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE")
YELLOW_FILL = PatternFill("solid", start_color="FFEB9C")
RED_FILL    = PatternFill("solid", start_color="FFC7CE")


def build_summary_sheet(ws, data: dict):
    headers = [
        "Address", "Found On", "Found At", "Age (days)",
        "Score", "Verdict", "Path",
        "Win Rate %", "Unique Tokens", "Total Trades",
        "Cost (ETH)", "P&L (ETH)", "ROI %",
        "Avg P&L/Trade (ETH)", "Is Bot"
    ]
    ws.row_dimensions[1].height = 20
    for col, h in enumerate(headers, 1):
        header_cell(ws.cell(row=1, column=col), h)

    for row_idx, (addr, entry) in enumerate(data.items(), 2):
        p   = entry.get("profile", {})
        s   = entry.get("score", {})
        alt = PatternFill("solid", start_color="EEF2F7") if row_idx % 2 == 0 else None

        score_val = s.get("total", 0)
        score_fill = GREEN_FILL if score_val >= 65 else YELLOW_FILL if score_val >= 50 else alt

        pnl_val  = p.get("total_pnl_eth", 0)
        pnl_fill = GREEN_FILL if pnl_val > 0 else RED_FILL if pnl_val < 0 else alt

        rows = [
            (1,  addr,                              None,      alt,        False),
            (2,  entry.get("found_on", ""),         None,      alt,        False),
            (3,  entry.get("found_at", "")[:10],    None,      alt,        False),
            (4,  p.get("age_days", 0),              None,      alt,        False),
            (5,  score_val,                         None,      score_fill, True),
            (6,  s.get("verdict", ""),              None,      alt,        False),
            (7,  s.get("path", ""),                 None,      alt,        False),
            (8,  p.get("win_rate", 0),              "0.0",     alt,        False),
            (9,  p.get("unique_tokens", 0),         None,      alt,        False),
            (10, p.get("total_trades", 0),          None,      alt,        False),
            (11, p.get("total_cost_eth", 0),        "0.0000",  alt,        False),
            (12, pnl_val,                           "0.0000",  pnl_fill,   True),
            (13, p.get("roi_pct", 0),               "0.00",    alt,        False),
            (14, p.get("avg_pnl_per_trade", 0),     "0.0000",  alt,        False),
            (15, "Yes" if p.get("is_bot") else "No", None,     alt,        False),
        ]
        for col, val, fmt, fill, bold in rows:
            body_cell(ws.cell(row=row_idx, column=col), val, fmt=fmt, fill=fill, bold=bold)

    col_widths = [44, 10, 12, 10, 7, 12, 20, 10, 14, 12, 12, 12, 10, 18, 8]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_trades_sheet(ws, data: dict, activity_key: str):
    headers = ["Wallet", "Action", "Token Symbol", "Token Address", "ETH Amount", "Timestamp", "Tx Hash"]
    ws.row_dimensions[1].height = 20
    for col, h in enumerate(headers, 1):
        header_cell(ws.cell(row=1, column=col), h)

    row_idx = 2
    for addr, entry in data.items():
        if activity_key == "trade_history":
            trades = entry.get("profile", {}).get("trade_history", [])
        else:
            trades = entry.get("activity", [])

        for trade in trades:
            action = trade.get("action", "").upper()
            fill   = GREEN_FILL if action == "BUY" else RED_FILL

            vals = [
                addr,
                action,
                trade.get("token_symbol", ""),
                trade.get("token_address", ""),
                trade.get("eth_amount", 0),
                trade.get("timestamp", "")[:19].replace("T", " "),
                trade.get("tx_hash", ""),
            ]
            for col, val in enumerate(vals, 1):
                fmt = "0.0000" if col == 5 else None
                body_cell(ws.cell(row=row_idx, column=col), val, fmt=fmt, fill=fill)
            row_idx += 1

    col_widths = [44, 8, 14, 44, 12, 20, 68]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    return row_idx - 2


def export(watchlist_file=WATCHLIST_FILE, output_file=OUTPUT_FILE):
    data = load_watchlist()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Wallet Summary"
    build_summary_sheet(ws1, data)

    ws2 = wb.create_sheet("Trade History")
    trade_count = build_trades_sheet(ws2, data, "trade_history")

    ws3 = wb.create_sheet("Recent Activity")
    activity_count = build_trades_sheet(ws3, data, "activity")

    wb.save(output_file)

    print(f"Exported: {output_file}")
    print(f"  {len(data)} wallet(s)")
    print(f"  {trade_count} trade history rows")
    print(f"  {activity_count} recent activity rows")
    print(f"  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == "__main__":
    export()
